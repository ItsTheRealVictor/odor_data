import openpyxl as opx
import pandas as pd

file_name = 'odor_scrubber1_raw.xlsx'
wkbk = opx.load_workbook(file_name)

sheet_names = wkbk.sheetnames

rows = []
for name in sheet_names:

    sheet = wkbk[name]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({
            'date/time': row[0],
            'column': row[1],
            'value': row[2],
            'unit': row[3]
        })


df = pd.DataFrame(rows)

# Sort the "date/time" column in ascending order
df['date/time'] = pd.to_datetime(df['date/time'])
df.sort_values(by='date/time', inplace=True)

# Perform your pivot and reset operations
pivoted_df = df.pivot_table(index='date/time', columns='column', values='value', aggfunc='first')
pivoted_df = pivoted_df.reset_index()

pivoted_df.to_excel(f'{file_name}_cleaned.xlsx', index=False)
print('finished')